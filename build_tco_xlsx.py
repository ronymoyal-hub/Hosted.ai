import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BLUE = Font(name='Arial', color='0000FF')
BLACK = Font(name='Arial', color='000000')
GREEN = Font(name='Arial', color='008000')
BOLD = Font(name='Arial', bold=True)
BOLD_WHITE = Font(name='Arial', bold=True, color='FFFFFF')
TITLE = Font(name='Arial', bold=True, size=14)
SECTION_FILL = PatternFill('solid', start_color='1F2937')
HEADER_FILL = PatternFill('solid', start_color='0B5394')
CHECK_FILL_OK = PatternFill('solid', start_color='D9EAD3')
CHECK_FILL_BAD = PatternFill('solid', start_color='F4CCCC')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
USD3 = '$#,##0.000;($#,##0.000);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM0 = '#,##0;(#,##0);"-"'

wb = Workbook()

def style_section(ws, row, col, text, span=4):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD_WHITE
    c.fill = SECTION_FILL
    for cc in range(col, col+span):
        ws.cell(row=row, column=cc).fill = SECTION_FILL

def label(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD if bold else Font(name='Arial')
    return c

def val(ws, row, col, formula_or_value, font=BLACK, numfmt=None, comment=None):
    c = ws.cell(row=row, column=col, value=formula_or_value)
    c.font = font
    if numfmt:
        c.number_format = numfmt
    if comment:
        c.comment = Comment(comment, 'Source')
    return c

# =========================================================================
# SHEET 1: COVER
# =========================================================================
cover = wb.active
cover.title = 'Cover'
cover.sheet_view.showGridLines = False
cover.column_dimensions['A'].width = 3
cover.column_dimensions['B'].width = 100
cover['B2'] = 'AI Cloud TCO & Returns Model'
cover['B2'].font = TITLE
cover['B4'] = 'Independent build for illustrative/planning use for the benefit of Hosted.ai customers.'
cover['B4'].alignment = Alignment(wrap_text=True, vertical='top')
cover.row_dimensions[4].height = 60
cover['B6'] = 'What this workbook contains:'
cover['B6'].font = BOLD
items = [
    '1. Assumptions — every input driving the model, organized by topic. Blue = input you can change.',
    '2. TCO Build-Up — the $/GPU-hr cost of ownership, split into capex recovery, power/colocation, maintenance, insurance.',
    '3. Income Statement — 5-year P&L (revenue through net income).',
    '4. Balance Sheet — 5-year (+ Year 0) balance sheet including a built-in balance check.',
    '5. Cash Flow — 5-year cash flow statement (operating / investing / financing).',
    '6. Returns — unlevered project and levered equity cash flows, IRR, NPV, and residual value.',
    '7. Hosted.ai ROI — always-on Without vs. With Hosted.ai comparison: overcommit-driven revenue uplift, '
    'the VRAM-based license fee, and the resulting IRR/NPV delta.',
]
r = 7
for it in items:
    cover.cell(row=r, column=2, value=it).alignment = Alignment(wrap_text=True)
    r += 1
r += 1
cover.cell(row=r, column=2, value='Methodology notes:').font = BOLD
r += 1
notes = [
    'TCO/hr = capex annualized via a capital-recovery factor at the project WACC, plus power/colocation, maintenance, '
    'and insurance, all divided by annual GPU-hours delivered at the assumed utilization rate.',
    'Revenue = GPU count x 8,760 hours x utilization x rental price ($/GPU-hr), with price declining annually.',
    'Depreciation is straight-line to salvage value over the stated useful life.',
    'Debt amortizes in equal annual principal installments over the debt term.',
    'Project FCF = EBIT x (1-tax) + D&A - capex. Equity FCF = Net Income + D&A - debt principal repaid.',
    'A residual/salvage value is added to the final projection year for both project and equity cash flows.',
    'This build uses a fixed 5-year explicit forecast; extend columns manually for a longer horizon.',
]
for n in notes:
    cover.cell(row=r, column=2, value='- ' + n).alignment = Alignment(wrap_text=True)
    cover.row_dimensions[r].height = 28
    r += 1

# =========================================================================
# SHEET 2: ASSUMPTIONS
# =========================================================================
a = wb.create_sheet('Assumptions')
a.sheet_view.showGridLines = False
a.column_dimensions['A'].width = 2
a.column_dimensions['B'].width = 38
a.column_dimensions['C'].width = 16
a.column_dimensions['D'].width = 22
a.column_dimensions['E'].width = 46

a['B1'] = 'Assumptions'
a['B1'].font = TITLE
a['B2'] = 'Blue = hardcoded input. Black = calculated. Edit blue cells to run scenarios.'
a['B2'].font = Font(name='Arial', italic=True, color='666666')

style_section(a, 4, 2, '1. GPU & CLUSTER')
label(a,5,2,'GPU / system type'); val(a,5,3,'H100', BLUE)
label(a,6,2,'Number of GPUs'); val(a,6,3,8192, BLUE, NUM0)
label(a,7,2,'All-in capex per GPU ($)'); val(a,7,3,33000, BLUE, USD0,
      'Fully-loaded 8-GPU HGX H100 system incl. NVLink fabric, networking, storage, integration; ~$33k/GPU. '
      'Market research synthesis, 2026 (getdeploying.com, mercatus-ai.com, cloudzero.com H100 pricing surveys).')
label(a,8,2,'IT power draw per GPU (W, all-in)'); val(a,8,3,1300, BLUE, NUM0,
      'H100 SXM GPU TDP ~700W; ~1,300W all-in incl. host CPU/DRAM/NIC/storage share.')

style_section(a, 10, 2, '2. POWER & FACILITY COST')
label(a,11,2,'Power mode (1=Colocation, 0=Owned facility)'); val(a,11,3,1, BLUE, NUM0)
label(a,12,2,'Colocation rate ($/kW-month, all-in)'); val(a,12,3,180, BLUE, USD0,
      'North American primary-market colo average for 250-500kW deals was ~$196/kW-mo in H2 2025 (CBRE); '
      'large multi-MW GPU cluster deals typically clear somewhat below the small-block retail rate.')
label(a,13,2,'Electricity rate ($/kWh, owned facility)'); val(a,13,3,0.08, BLUE, '$#,##0.000')
label(a,14,2,'Facility PUE (owned facility)'); val(a,14,3,1.3, BLUE, '0.00')

style_section(a, 16, 2, '3. FINANCING')
label(a,17,2,'Debt % of capex'); val(a,17,3,0.50, BLUE, PCT1,
      'Neoclouds typically finance 50-80% of GPU capex with debt collateralized against contracted cash flows '
      '(e.g. CoreWeave DDTL structures).')
label(a,18,2,'All-in debt interest rate'); val(a,18,3,0.08, BLUE, PCT1,
      "CoreWeave's investment-grade GPU-backed DDTL 4.0 priced at SOFR+2.25%; blended neocloud debt stacks run ~7-10% all-in.")
label(a,19,2,'Debt term (years)'); val(a,19,3,5, BLUE, NUM0)

style_section(a, 21, 2, '4. DEPRECIATION & USEFUL LIFE')
label(a,22,2,'Accounting/economic useful life (years)'); val(a,22,3,5, BLUE, NUM0,
      'Hyperscalers/neoclouds typically book 4-6yr GPU server depreciation. Bear case (M. Burry, Nov-2025) argues '
      'economic life is closer to 2-3yrs — stress test by lowering this input.')
label(a,23,2,'Residual / salvage value (% of capex)'); val(a,23,3,0.10, BLUE, PCT1)

style_section(a, 25, 2, '5. REVENUE & PRICING')
label(a,26,2,'Rental price, Year 1 ($/GPU-hr)'); val(a,26,3,2.50, BLUE, USD2,
      '2026 market range for H100: hyperscalers $2-11/hr, specialized neoclouds $1.50-2.50/hr, CoreWeave on-demand '
      '~$6.16/GPU-hr normalized from 8-GPU HGX node pricing. $2.50 reflects a competitive neocloud on-demand rate.')
label(a,27,2,'Steady-state utilization'); val(a,27,3,0.80, BLUE, PCT1)
label(a,28,2,'Year-1 ramp (% of steady-state utilization)'); val(a,28,3,0.80, BLUE, PCT1)
label(a,29,2,'Annual rental price decline'); val(a,29,3,0.05, BLUE, PCT1,
      'GPU $/hr pricing has historically declined as newer accelerators launch and supply catches up with demand.')

style_section(a, 31, 2, '6. OPEX & TAX')
label(a,32,2,'Maintenance / support (% of capex, per yr)'); val(a,32,3,0.03, BLUE, PCT1)
label(a,33,2,'Insurance & other (% of capex, per yr)'); val(a,33,3,0.01, BLUE, PCT1)
label(a,34,2,'SG&A (% of revenue)'); val(a,34,3,0.08, BLUE, PCT1)
label(a,35,2,'Effective tax rate'); val(a,35,3,0.21, BLUE, PCT1)

style_section(a, 37, 2, '7. HORIZON & DISCOUNT RATES')
label(a,38,2,'Analysis horizon (years, fixed build)'); val(a,38,3,5, BLUE, NUM0)
label(a,39,2,'WACC / project discount rate'); val(a,39,3,0.12, BLUE, PCT1)
label(a,40,2,'Cost of equity / equity discount rate'); val(a,40,3,0.18, BLUE, PCT1)

style_section(a, 42, 2, 'DERIVED (calculated — do not overwrite)')
label(a,43,2,'Total capex ($)'); val(a,43,3,'=C6*C7', BLACK, USD0)
label(a,44,2,'Debt amount ($)'); val(a,44,3,'=C43*C17', BLACK, USD0)
label(a,45,2,'Equity amount ($)'); val(a,45,3,'=C43-C44', BLACK, USD0)
label(a,46,2,'Total IT load (kW)'); val(a,46,3,'=C6*C8/1000', BLACK, NUM0)
label(a,47,2,'Annual power/facility cost ($)'); val(a,47,3,'=IF(C11=1,C46*C12*12,C46*C14*8760*C13)', BLACK, USD0)
label(a,48,2,'Capital recovery factor (CRF)'); val(a,48,3,'=C39*(1+C39)^C22/((1+C39)^C22-1)', BLACK, '0.0000')
label(a,49,2,'Annual capex charge — capital recovery ($)'); val(a,49,3,'=C43*(1-C23)*C48+C43*C23*C39', BLACK, USD0)
label(a,50,2,'Annual maintenance ($)'); val(a,50,3,'=C43*C32', BLACK, USD0)
label(a,51,2,'Annual insurance ($)'); val(a,51,3,'=C43*C33', BLACK, USD0)
label(a,52,2,'Annual GPU-hours at utilization'); val(a,52,3,'=C6*8760*C27', BLACK, NUM0)
label(a,53,2,'Straight-line depreciation ($/yr)'); val(a,53,3,'=(C43-C43*C23)/C22', BLACK, USD0)
label(a,54,2,'Debt principal repayment ($/yr)'); val(a,54,3,'=C44/C19', BLACK, USD0)

style_section(a, 56, 2, 'REFERENCE — GPU PRESET LIBRARY (for context; not linked into calc)')
hdrs = ['GPU / System','Capex/GPU ($)','IT Power/GPU (W)','Illustrative rental price ($/GPU-hr)']
for i,h in enumerate(hdrs):
    c = a.cell(row=57, column=2+i, value=h); c.font = BOLD
presets = [
    ('NVIDIA H100 (HGX 8-GPU)', 33000, 1300, 2.50),
    ('NVIDIA H200 (HGX 8-GPU)', 40000, 1400, 3.00),
    ('NVIDIA B200 (air-cooled HGX)', 50000, 1700, 4.50),
    ('NVIDIA GB200 NVL72 (rack-scale)', 54167, 1833, 18.00),
]
rr = 58
for name, capex, power, price in presets:
    a.cell(row=rr, column=2, value=name)
    a.cell(row=rr, column=3, value=capex).number_format = USD0
    a.cell(row=rr, column=4, value=power).number_format = NUM0
    a.cell(row=rr, column=5, value=price).number_format = USD2
    rr += 1
a.cell(row=rr+1, column=2, value=('Sources: 8-GPU HGX H100 systems ~$250k-$320k (2026 OEM survey); GB200 NVL72 rack '
       '~$3.0-3.4M compute + ~$3.9M all-in / 72 GPUs; B200 discrete GPU $45-55k; GB200 rack power 132kW/72 GPUs (NVIDIA, '
       'HPE, Spheron Network specs, 2026).')).font = Font(name='Arial', italic=True, size=9, color='666666')
a.cell(row=rr+1, column=2).alignment = Alignment(wrap_text=True)
a.row_dimensions[rr+1].height = 30

# ---- 8. HOSTED.AI PLATFORM (appended section; referenced by TCO/Income Statement/Returns/Hosted_ai_ROI) ----
hai_start = rr + 3
style_section(a, hai_start, 2, '8. HOSTED.AI PLATFORM')
r8 = hai_start + 1
label(a,r8,2,'Include Hosted.ai in main statements (1=Yes, 0=No)'); val(a,r8,3,1, BLUE, NUM0); ROW_HAI_TOGGLE=r8; r8+=1
label(a,r8,2,'VRAM per GPU (GB)'); val(a,r8,3,80, BLUE, NUM0,
      'H100=80GB, H200=141GB, B200=180GB, GB200 NVL72=188GB. Used to size the Hosted.ai VRAM-based license fee.'); ROW_VRAM=r8; r8+=1
label(a,r8,2,'GPU overcommit / multi-tenancy ratio (x)'); val(a,r8,3,3.0, BLUE, '0.0"x"',
      "Hosted.ai's multi-tenant pooling lets you sell more vGPU-hours than physical GPU-hours exist (2x-10x per "
      "hosted.ai/neoclouds; 3x matches internal CSP reference modeling)."); ROW_OC=r8; r8+=1
label(a,r8,2,'Virtualized service price premium (%)'); val(a,r8,3,0.05, BLUE, PCT1,
      'Managed multi-tenant service commands a premium over raw bare-metal market pricing.'); ROW_MARKUP=r8; r8+=1
label(a,r8,2,'HAI allocation fee ($/GB-hr)'); val(a,r8,3,0.0005, BLUE, '$#,##0.0000',
      'hosted.ai/pricing: VRAM managed rate, $0.0005/GB-h.'); ROW_ALLOC=r8; r8+=1
label(a,r8,2,'HAI utilization fee ($/GB-hr)'); val(a,r8,3,0.0030, BLUE, '$#,##0.0000',
      'hosted.ai/pricing: VRAM consumed rate, $0.0030/GB-h.'); ROW_UTILFEE=r8; r8+=1
label(a,r8,2,'HAI minimum monthly fee ($)'); val(a,r8,3,750, BLUE, USD0,
      'hosted.ai/pricing: stated minimum $750/month.'); ROW_MIN=r8; r8+=1

r8 += 1
style_section(a, r8, 2, 'HOSTED.AI — DERIVED (calculated — do not overwrite)'); r8+=1
label(a,r8,2,'Total VRAM managed (GB)'); val(a,r8,3,f'=C{6}*C{ROW_VRAM}', BLACK, NUM0); ROW_VRAM_TOTAL=r8; r8+=1
label(a,r8,2,'Effective overcommit ratio (applied)'); val(a,r8,3,f'=IF(C{ROW_HAI_TOGGLE}=1,C{ROW_OC},1)', BLACK, '0.0"x"'); ROW_EFF_OC=r8; r8+=1
label(a,r8,2,'Effective price premium (applied)'); val(a,r8,3,f'=IF(C{ROW_HAI_TOGGLE}=1,C{ROW_MARKUP},0)', BLACK, PCT1); ROW_EFF_MARKUP=r8; r8+=1
label(a,r8,2,'Effective utilization (applied, same both scenarios)'); val(a,r8,3,'=C27', BLACK, PCT1); ROW_EFF_UTIL=r8; r8+=1
label(a,r8,2,'Effective rental price, Year 1 ($/GPU-hr)'); val(a,r8,3,f'=C26*(1+C{ROW_EFF_MARKUP})', BLACK, USD2); ROW_EFF_PRICE=r8; r8+=1
label(a,r8,2,'Monthly Hosted.ai fee @ steady-state util. ($)'); val(a,r8,3,
      f'=IF(C{ROW_HAI_TOGGLE}=1,MAX(C{ROW_MIN},C{ROW_VRAM_TOTAL}*720*C{ROW_ALLOC}+C{ROW_VRAM_TOTAL}*720*C{ROW_EFF_UTIL}*C{ROW_UTILFEE}),0)',
      BLACK, USD0); ROW_MONTHLY_FEE=r8; r8+=1
label(a,r8,2,'Annual Hosted.ai fee @ steady-state ($)'); val(a,r8,3,f'=C{ROW_MONTHLY_FEE}*12', BLACK, USD0); ROW_ANNUAL_FEE=r8; r8+=1
label(a,r8,2,'Annual billable GPU-hours (incl. overcommit)'); val(a,r8,3,f'=C6*8760*C{ROW_EFF_UTIL}*C{ROW_EFF_OC}', BLACK, NUM0); ROW_BILLABLE_HRS=r8; r8+=1

hai_end = r8 - 1
for row in a.iter_rows(min_row=4, max_row=hai_end, min_col=2, max_col=5):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 3: TCO BUILD-UP
# =========================================================================
t = wb.create_sheet('TCO_BuildUp')
t.sheet_view.showGridLines = False
t.column_dimensions['A'].width = 2
t.column_dimensions['B'].width = 34
for col in 'CDEF':
    t.column_dimensions[col].width = 18

t['B1'] = 'TCO Build-Up ($ / GPU-hr)'
t['B1'].font = TITLE

hdr_row = 3
for i, h in enumerate(['Component', '$ / GPU-hr', '% of TCO', 'Annualized ($)']):
    c = t.cell(row=hdr_row, column=2+i, value=h)
    c.font = BOLD_WHITE; c.fill = HEADER_FILL

BILLHRS = f'Assumptions!C{ROW_BILLABLE_HRS}'
rows = [
    ('Capex — capital recovery @ WACC', f'=Assumptions!C49/{BILLHRS}', '=Assumptions!C49'),
    ('Power / colocation',               f'=Assumptions!C47/{BILLHRS}', '=Assumptions!C47'),
    ('Maintenance & support',            f'=Assumptions!C50/{BILLHRS}', '=Assumptions!C50'),
    ('Insurance & other',                f'=Assumptions!C51/{BILLHRS}', '=Assumptions!C51'),
    ('Hosted.ai license fee',            f'=Assumptions!C{ROW_ANNUAL_FEE}/{BILLHRS}', f'=Assumptions!C{ROW_ANNUAL_FEE}'),
]
r = 4
first_r = r
for name, perhr_formula, annual_formula in rows:
    label(t, r, 2, name)
    val(t, r, 3, perhr_formula, GREEN, USD3)
    annual_cell_ref = f'D{r}'
    val(t, r, 5, annual_formula, GREEN, USD0)
    r += 1
last_r = r - 1
label(t, r, 2, 'Total TCO', bold=True)
c = val(t, r, 3, f'=SUM(C{first_r}:C{last_r})', BLACK, USD3); c.font = BOLD
val(t, r, 5, f'=SUM(E{first_r}:E{last_r})', BLACK, USD0).font = BOLD
total_row = r
for rr_ in range(first_r, total_row+1):
    t.cell(row=rr_, column=4, value=f'=C{rr_}/$C${total_row}').number_format = PCT1

r += 2
label(t, r, 2, 'Effective rental price, Year 1 ($/GPU-hr)')
val(t, r, 3, f'=Assumptions!C{ROW_EFF_PRICE}', GREEN, USD2)
price_row = r
r += 1
label(t, r, 2, 'Gross margin @ list price', bold=True)
val(t, r, 3, f'=(C{price_row}-C{total_row})/C{price_row}', BLACK, PCT1).font = BOLD
r += 1
label(t, r, 2, 'Billable GPU-hours/yr (incl. overcommit)')
val(t, r, 3, f'=Assumptions!C{ROW_BILLABLE_HRS}', GREEN, NUM0)

for row in t.iter_rows(min_row=3, max_row=r, min_col=2, max_col=5):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 4: INCOME STATEMENT  (columns C..G = Year1..Year5)
# =========================================================================
inc = wb.create_sheet('Income_Statement')
inc.sheet_view.showGridLines = False
inc.column_dimensions['A'].width = 2
inc.column_dimensions['B'].width = 32
for col in 'CDEFG':
    inc.column_dimensions[col].width = 16

inc['B1'] = 'Income Statement'
inc['B1'].font = TITLE

yr_cols = ['C','D','E','F','G']
for i, col in enumerate(yr_cols):
    c = inc[f'{col}3']; c.value = f'Year {i+1}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL
inc['B3'] = ''; inc['B3'].fill = HEADER_FILL

def set_row(ws, row, label_text, formulas, font=BLACK, numfmt=USD0, bold=False):
    lbl = label(ws, row, 2, label_text, bold=bold)
    for col, f in zip(yr_cols, formulas):
        cell = ws[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, color=font.color.rgb if hasattr(font,'color') and font.color else '000000')
        cell.number_format = numfmt

# Row 4: utilization (effective — reflects With/Without Hosted.ai toggle)
set_row(inc, 4, 'Utilization (effective)', [
    f'=Assumptions!$C${ROW_EFF_UTIL}*Assumptions!$C$28',
    f'=Assumptions!$C${ROW_EFF_UTIL}', f'=Assumptions!$C${ROW_EFF_UTIL}', f'=Assumptions!$C${ROW_EFF_UTIL}', f'=Assumptions!$C${ROW_EFF_UTIL}'
], GREEN, PCT1)
# Row 5: price (effective — includes Hosted.ai virtualized premium if enabled)
set_row(inc, 5, 'Rental price, effective ($/GPU-hr)', [
    f'=Assumptions!$C${ROW_EFF_PRICE}',
    '=C5*(1-Assumptions!$C$29)', '=D5*(1-Assumptions!$C$29)', '=E5*(1-Assumptions!$C$29)', '=F5*(1-Assumptions!$C$29)'
], GREEN, USD2)
# Row 6: revenue (includes overcommit multiplier on billable GPU-hours)
set_row(inc, 6, 'Revenue', [
    f'={col}4*{col}5*Assumptions!$C$6*8760*Assumptions!$C${ROW_EFF_OC}' for col in yr_cols
], BLACK, USD0, bold=False)
# Row 7: COGS (power/colo + maint. + insurance + Hosted.ai license fee, usage-based on that year's utilization)
hai_fee_formula = (f'IF(Assumptions!$C${ROW_HAI_TOGGLE}=1,'
                   f'MAX(Assumptions!$C${ROW_MIN},Assumptions!$C${ROW_VRAM_TOTAL}*720*Assumptions!$C${ROW_ALLOC}'
                   f'+Assumptions!$C${ROW_VRAM_TOTAL}*720*{{col}}4*Assumptions!$C${ROW_UTILFEE})*12,0)')
set_row(inc, 7, 'COGS: power/colo+maint+insur+Hosted.ai', [
    f'=Assumptions!$C$47+Assumptions!$C$50+Assumptions!$C$51+{hai_fee_formula.format(col=col)}' for col in yr_cols
], GREEN, USD0)
# Row 7b: memo — Hosted.ai fee portion only, broken out for visibility
set_row(inc, 17, 'memo: Hosted.ai license fee (incl. above)', [
    f'={hai_fee_formula.format(col=col)}' for col in yr_cols
], GREEN, USD0)
inc['B17'].font = Font(name='Arial', italic=True, color='666666')
for col in yr_cols:
    inc[f'{col}17'].font = Font(name='Arial', italic=True, color='008000')
# Row 8: gross profit
set_row(inc, 8, 'Gross Profit', [f'={col}6-{col}7' for col in yr_cols], BLACK, USD0, bold=True)
# Row 9: SG&A
set_row(inc, 9, 'SG&A', [f'={col}6*Assumptions!$C$34' for col in yr_cols], GREEN, USD0)
# Row 10: EBITDA
set_row(inc, 10, 'EBITDA', [f'={col}8-{col}9' for col in yr_cols], BLACK, USD0, bold=True)
# Row 11: D&A (capped cumulative)
dep_formulas = []
base = 'Assumptions!$C$43*(1-Assumptions!$C$23)'
dep_formulas.append(f'=MIN(Assumptions!$C$53,{base})')
dep_formulas.append(f'=MIN(Assumptions!$C$53,{base}-C11)')
dep_formulas.append(f'=MIN(Assumptions!$C$53,{base}-C11-D11)')
dep_formulas.append(f'=MIN(Assumptions!$C$53,{base}-C11-D11-E11)')
dep_formulas.append(f'=MIN(Assumptions!$C$53,{base}-C11-D11-E11-F11)')
set_row(inc, 11, 'Depreciation & Amortization', dep_formulas, GREEN, USD0)
# Row 12: EBIT
set_row(inc, 12, 'EBIT', [f'={col}10-{col}11' for col in yr_cols], BLACK, USD0, bold=True)
# Row 13: interest (uses Balance_Sheet prior column debt balance, row 11)
bs_prior_cols = ['B','C','D','E','F']  # BS Year0..Year4 debt balances feed IS Yr1..Yr5 interest
set_row(inc, 13, 'Interest Expense', [
    f'=Balance_Sheet!{bcol}11*Assumptions!$C$18' for bcol in bs_prior_cols
], GREEN, USD0)
# Row 14: EBT
set_row(inc, 14, 'EBT', [f'={col}12-{col}13' for col in yr_cols], BLACK, USD0)
# Row 15: tax
set_row(inc, 15, 'Tax', [f'=MAX(0,{col}14)*Assumptions!$C$35' for col in yr_cols], GREEN, USD0)
# Row 16: net income
set_row(inc, 16, 'Net Income', [f'={col}14-{col}15' for col in yr_cols], BLACK, USD0, bold=True)

for row in inc.iter_rows(min_row=3, max_row=17, min_col=2, max_col=7):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 5: BALANCE SHEET (columns B..G = Year0..Year5)
# =========================================================================
bs = wb.create_sheet('Balance_Sheet')
bs.sheet_view.showGridLines = False
bs.column_dimensions['A'].width = 2
bs.column_dimensions['B'].width = 30
for col in 'CDEFG':
    bs.column_dimensions[col].width = 16
bsy_cols = ['B','C','D','E','F','G']

bs['B1'] = 'Balance Sheet'
bs['B1'].font = TITLE
for i, col in enumerate(bsy_cols):
    c = bs[f'{col}3']; c.value = f'Year {i}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL

style_section(bs, 5, 2, 'ASSETS', span=6)

def set_bs_row(ws, row, label_text, formulas, font=BLACK, numfmt=USD0, bold=False):
    label(ws, row, 2, label_text, bold=bold)
    for col, f in zip(bsy_cols, formulas):
        cell = ws[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, color=(font.color.rgb if font.color else '000000'))
        cell.number_format = numfmt

# Row 6: Cash. Year0 = 0 (blue input). Yr1..5 = prior + net change in cash from Cash_Flow sheet.
set_bs_row(bs, 6, 'Cash', [
    0,
    '=B6+Cash_Flow!C9', '=C6+Cash_Flow!D9', '=D6+Cash_Flow!E9', '=E6+Cash_Flow!F9', '=F6+Cash_Flow!G9'
])
bs['B6'].font = BLUE
# Row 7: Net PP&E. Year0 = total capex. Yr t = prior - D&A(t) from Income_Statement row 11
set_bs_row(bs, 7, 'Net PP&E', [
    '=Assumptions!C43',
    '=B7-Income_Statement!C11', '=C7-Income_Statement!D11', '=D7-Income_Statement!E11',
    '=E7-Income_Statement!F11', '=F7-Income_Statement!G11'
], GREEN)
# Row 8: Total assets
set_bs_row(bs, 8, 'Total Assets', [f'=SUM({c}6:{c}7)' for c in bsy_cols], BLACK, USD0, bold=True)

style_section(bs, 10, 2, 'LIABILITIES & EQUITY', span=6)
# Row 11: Debt. Year0 = debt amount. Yr t = prior - principal repayment (if within debt term)
debt_formulas = ['=Assumptions!C44']
for i, col in enumerate(bsy_cols[1:]):  # i = 0..4 -> year 1..5
    prior_col = bsy_cols[i]
    yr_num = i + 1
    debt_formulas.append(f'={prior_col}11-IF({yr_num}<=Assumptions!$C$19,Assumptions!$C$54,0)')
set_bs_row(bs, 11, 'Debt', debt_formulas, GREEN)
# Row 12: Paid-in equity (constant)
set_bs_row(bs, 12, 'Paid-in Equity', ['=Assumptions!C45','=B12','=C12','=D12','=E12','=F12'], GREEN)
# Row 13: Retained earnings. Year0 = 0. Yr t = prior + net income(t)
set_bs_row(bs, 13, 'Retained Earnings', [
    0,
    '=B13+Income_Statement!C16', '=C13+Income_Statement!D16', '=D13+Income_Statement!E16',
    '=E13+Income_Statement!F16', '=F13+Income_Statement!G16'
], GREEN)
bs['B13'].font = BLUE
# Row 14: total liab + equity
set_bs_row(bs, 14, 'Total Liabilities + Equity', [f'=SUM({c}11:{c}13)' for c in bsy_cols], BLACK, USD0, bold=True)

# Row 16: balance check
label(bs, 16, 2, 'Balance Check (Assets - Liab&Equity)', bold=True)
for col in bsy_cols:
    cell = bs[f'{col}16']
    cell.value = f'={col}8-{col}14'
    cell.number_format = USD0
    cell.font = BOLD

for row in bs.iter_rows(min_row=3, max_row=16, min_col=2, max_col=7):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 6: CASH FLOW (columns C..G = Year1..Year5)
# =========================================================================
cf = wb.create_sheet('Cash_Flow')
cf.sheet_view.showGridLines = False
cf.column_dimensions['A'].width = 2
cf.column_dimensions['B'].width = 32
for col in 'CDEFG':
    cf.column_dimensions[col].width = 16

cf['B1'] = 'Cash Flow Statement'
cf['B1'].font = TITLE
for i, col in enumerate(yr_cols):
    c = cf[f'{col}3']; c.value = f'Year {i+1}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL

def set_cf_row(ws, row, label_text, formulas, font=BLACK, numfmt=USD0, bold=False):
    label(ws, row, 2, label_text, bold=bold)
    for col, f in zip(yr_cols, formulas):
        cell = ws[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, color=(font.color.rgb if font.color else '000000'))
        cell.number_format = numfmt

set_cf_row(cf, 4, 'Net Income', [f'=Income_Statement!{c}16' for c in yr_cols], GREEN)
set_cf_row(cf, 5, '+ D&A', [f'=Income_Statement!{c}11' for c in yr_cols], GREEN)
set_cf_row(cf, 6, 'Cash Flow from Operations', [f'={c}4+{c}5' for c in yr_cols], BLACK, USD0, bold=True)
set_cf_row(cf, 7, 'Capex (Investing)', [0,0,0,0,0])
cf['C7'].font = BLUE; cf['D7'].font=BLUE; cf['E7'].font=BLUE; cf['F7'].font=BLUE; cf['G7'].font=BLUE
principal_formulas = []
for i in range(5):
    yr_num = i + 1
    principal_formulas.append(f'=-IF({yr_num}<=Assumptions!$C$19,Assumptions!$C$54,0)')
set_cf_row(cf, 8, 'Debt Principal Repayment (Financing)', principal_formulas, GREEN)
set_cf_row(cf, 9, 'Net Change in Cash', [f'={c}6+{c}7+{c}8' for c in yr_cols], BLACK, USD0, bold=True)
set_cf_row(cf, 10, 'Ending Cash Balance (tie to Balance Sheet)', [f'=Balance_Sheet!{c}6' for c in bsy_cols[1:]], GREEN)

for row in cf.iter_rows(min_row=3, max_row=10, min_col=2, max_col=7):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 7: RETURNS (columns B..G = Year0..Year5)
# =========================================================================
ret = wb.create_sheet('Returns')
ret.sheet_view.showGridLines = False
ret.column_dimensions['A'].width = 2
ret.column_dimensions['B'].width = 34
for col in 'CDEFG':
    ret.column_dimensions[col].width = 16

ret['B1'] = 'Returns — Project & Equity Cash Flow'
ret['B1'].font = TITLE
ret_cols = ['B','C','D','E','F','G']
for i, col in enumerate(ret_cols):
    c = ret[f'{col}3']; c.value = f'Year {i}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL

def set_ret_row(ws, row, label_text, formulas, font=BLACK, numfmt=USD0, bold=False):
    label(ws, row, 2, label_text, bold=bold)
    for col, f in zip(ret_cols, formulas):
        cell = ws[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, color=(font.color.rgb if font.color else '000000'))
        cell.number_format = numfmt

# Project FCF: Year0 = -total capex. Yr1-4 = EBIT*(1-tax)+D&A-capex(investing=0).
# Yr5 additionally adds residual/salvage value.
proj_formulas = ['=-Assumptions!C43']
is_map = {'C':'C','D':'D','E':'E','F':'F','G':'G'}
for i, col in enumerate(yr_cols):
    yr_num = i + 1
    base_formula = f'Income_Statement!{col}12*(1-Assumptions!$C$35)+Income_Statement!{col}11-Cash_Flow!{col}7'
    if yr_num == 5:
        proj_formulas.append(f'={base_formula}+Assumptions!$C$43*Assumptions!$C$23')
    else:
        proj_formulas.append(f'={base_formula}')
set_ret_row(ret, 4, 'Project FCF (unlevered)', proj_formulas, GREEN)

# Equity FCF: Year0 = -equity amount. Yr t = NI + D&A + CF!row8 (already negative). Yr5 add residual - ending debt.
eq_formulas = ['=-Assumptions!C45']
for i, col in enumerate(yr_cols):
    yr_num = i + 1
    base_formula = f'Income_Statement!{col}16+Income_Statement!{col}11+Cash_Flow!{col}8'
    if yr_num == 5:
        eq_formulas.append(f'={base_formula}+Assumptions!$C$43*Assumptions!$C$23-Balance_Sheet!G11')
    else:
        eq_formulas.append(f'={base_formula}')
set_ret_row(ret, 5, 'Equity FCF (levered)', eq_formulas, GREEN)

set_ret_row(ret, 7, 'Cumulative Project FCF', [f'=SUM($B$4:{c}4)' for c in ret_cols], BLACK)
set_ret_row(ret, 8, 'Cumulative Equity FCF', [f'=SUM($B$5:{c}5)' for c in ret_cols], BLACK)

label(ret, 10, 2, 'Project IRR (unlevered)', bold=True)
ret['C10'] = '=IFERROR(IRR(B4:G4),"n/a")'; ret['C10'].number_format = PCT1; ret['C10'].font = Font(name='Arial', bold=True)
label(ret, 11, 2, 'Equity IRR (levered)', bold=True)
ret['C11'] = '=IFERROR(IRR(B5:G5),"n/a")'; ret['C11'].number_format = PCT1; ret['C11'].font = Font(name='Arial', bold=True)
label(ret, 12, 2, 'NPV — Project @ WACC', bold=True)
ret['C12'] = '=B4+NPV(Assumptions!C39,C4:G4)'; ret['C12'].number_format = USD0; ret['C12'].font = Font(name='Arial', bold=True)
label(ret, 13, 2, 'NPV — Equity @ Cost of Equity', bold=True)
ret['C13'] = '=B5+NPV(Assumptions!C40,C5:G5)'; ret['C13'].number_format = USD0; ret['C13'].font = Font(name='Arial', bold=True)
label(ret, 14, 2, 'Residual / salvage value at exit')
ret['C14'] = '=Assumptions!C43*Assumptions!C23'; ret['C14'].number_format = USD0
label(ret, 15, 2, 'Simple payback (project, break-even year)', bold=True)
ret['C15'] = '=IFERROR(MATCH(1,INDEX((C7:G7>=0)*1,0),0),"Beyond horizon")'
ret['C15'].font = Font(name='Arial', bold=True)

for row in ret.iter_rows(min_row=3, max_row=15, min_col=2, max_col=7):
    for cell in row:
        cell.border = BORDER

# =========================================================================
# SHEET 8: HOSTED.AI ROI (Without vs. With — always both, independent of the C66 toggle)
# =========================================================================
hai = wb.create_sheet('Hosted_ai_ROI')
hai.sheet_view.showGridLines = False
hai.column_dimensions['A'].width = 2
hai.column_dimensions['B'].width = 36
for col in 'CDEFG':
    hai.column_dimensions[col].width = 16

hai['B1'] = 'Hosted.ai ROI — Without vs. With Comparison'
hai['B1'].font = TITLE
hai['B2'] = ('Always computes both scenarios regardless of the Assumptions!C' + str(ROW_HAI_TOGGLE) +
             ' toggle. Capex, financing, D&A, and debt are identical in both — only revenue, pricing, and Hosted.ai fees differ.')
hai['B2'].font = Font(name='Arial', italic=True, color='666666')
hai.row_dimensions[2].height = 28
hai['B2'].alignment = Alignment(wrap_text=True)

for i, col in enumerate(yr_cols):
    c = hai[f'{col}3']; c.value = f'Year {i+1}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL

def set_hai_row(row, label_text, formulas, font=BLACK, numfmt=USD0, bold=False, italic=False):
    lbl = label(hai, row, 2, label_text, bold=bold)
    if italic: lbl.font = Font(name='Arial', italic=True, color='666666')
    for col, f in zip(yr_cols, formulas):
        cell = hai[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, italic=italic, color=(font.color.rgb if font.color else '000000'))
        cell.number_format = numfmt

r = 4
style_section(hai, r, 2, 'WITHOUT HOSTED.AI  (same utilization as With; no overcommit, no price premium, no license fee)', span=6); r+=1
row_w_util = r
set_hai_row(r, 'Utilization (held constant vs. With)', [
    '=Assumptions!$C$27*Assumptions!$C$28',
    '=Assumptions!$C$27', '=Assumptions!$C$27', '=Assumptions!$C$27', '=Assumptions!$C$27'
], GREEN, PCT1); r+=1
row_w_price = r
set_hai_row(r, 'Rental price ($/GPU-hr)', [
    '=Assumptions!$C$26', f'=C{row_w_price}*(1-Assumptions!$C$29)', f'=D{row_w_price}*(1-Assumptions!$C$29)',
    f'=E{row_w_price}*(1-Assumptions!$C$29)', f'=F{row_w_price}*(1-Assumptions!$C$29)'
], GREEN, USD2); r+=1
row_w_rev = r
set_hai_row(r, 'Revenue', [f'={col}{row_w_util}*{col}{row_w_price}*Assumptions!$C$6*8760' for col in yr_cols], BLACK, USD0); r+=1
row_w_cogs = r
set_hai_row(r, 'COGS (power/colo+maint+insurance)', ['=Assumptions!$C$47+Assumptions!$C$50+Assumptions!$C$51' for _ in yr_cols], GREEN, USD0); r+=1
row_w_gp = r
set_hai_row(r, 'Gross Profit', [f'={col}{row_w_rev}-{col}{row_w_cogs}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_w_sga = r
set_hai_row(r, 'SG&A', [f'={col}{row_w_rev}*Assumptions!$C$34' for col in yr_cols], GREEN, USD0); r+=1
row_w_ebitda = r
set_hai_row(r, 'EBITDA', [f'={col}{row_w_gp}-{col}{row_w_sga}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_w_da = r
set_hai_row(r, 'D&A (shared — unaffected by Hosted.ai)', [f'=Income_Statement!{col}11' for col in yr_cols], GREEN, USD0, italic=True); r+=1
row_w_ebit = r
set_hai_row(r, 'EBIT', [f'={col}{row_w_ebitda}-{col}{row_w_da}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_w_int = r
set_hai_row(r, 'Interest (shared — unaffected by Hosted.ai)', [f'=Income_Statement!{col}13' for col in yr_cols], GREEN, USD0, italic=True); r+=1
row_w_ebt = r
set_hai_row(r, 'EBT', [f'={col}{row_w_ebit}-{col}{row_w_int}' for col in yr_cols], BLACK, USD0); r+=1
row_w_tax = r
set_hai_row(r, 'Tax', [f'=MAX(0,{col}{row_w_ebt})*Assumptions!$C$35' for col in yr_cols], GREEN, USD0); r+=1
row_w_ni = r
set_hai_row(r, 'Net Income', [f'={col}{row_w_ebt}-{col}{row_w_tax}' for col in yr_cols], BLACK, USD0, bold=True); r+=1

r += 1
style_section(hai, r, 2, 'WITH HOSTED.AI  (overcommit + price premium + license fee, always on for this block)', span=6); r+=1
row_h_util = r
set_hai_row(r, 'Utilization', [
    '=Assumptions!$C$27*Assumptions!$C$28', '=Assumptions!$C$27', '=Assumptions!$C$27', '=Assumptions!$C$27', '=Assumptions!$C$27'
], GREEN, PCT1); r+=1
row_h_price = r
set_hai_row(r, 'Rental price, effective ($/GPU-hr)', [
    f'=Assumptions!$C$26*(1+Assumptions!$C${ROW_MARKUP})', f'=C{row_h_price}*(1-Assumptions!$C$29)', f'=D{row_h_price}*(1-Assumptions!$C$29)',
    f'=E{row_h_price}*(1-Assumptions!$C$29)', f'=F{row_h_price}*(1-Assumptions!$C$29)'
], GREEN, USD2); r+=1
row_h_rev = r
set_hai_row(r, 'Revenue (incl. overcommit)', [f'={col}{row_h_util}*{col}{row_h_price}*Assumptions!$C$6*8760*Assumptions!$C${ROW_OC}' for col in yr_cols], BLACK, USD0); r+=1
row_h_cogs = r
hai_fee_formula2 = (f'IF(TRUE,MAX(Assumptions!$C${ROW_MIN},Assumptions!$C${ROW_VRAM_TOTAL}*720*Assumptions!$C${ROW_ALLOC}'
                    f'+Assumptions!$C${ROW_VRAM_TOTAL}*720*{{col}}{row_h_util}*Assumptions!$C${ROW_UTILFEE})*12,0)')
set_hai_row(r, 'COGS: power/colo+maint+insur+Hosted.ai', [
    f'=Assumptions!$C$47+Assumptions!$C$50+Assumptions!$C$51+{hai_fee_formula2.format(col=col)}' for col in yr_cols
], GREEN, USD0); r+=1
row_h_gp = r
set_hai_row(r, 'Gross Profit', [f'={col}{row_h_rev}-{col}{row_h_cogs}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_h_sga = r
set_hai_row(r, 'SG&A', [f'={col}{row_h_rev}*Assumptions!$C$34' for col in yr_cols], GREEN, USD0); r+=1
row_h_ebitda = r
set_hai_row(r, 'EBITDA', [f'={col}{row_h_gp}-{col}{row_h_sga}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_h_da = r
set_hai_row(r, 'D&A (shared — unaffected by Hosted.ai)', [f'=Income_Statement!{col}11' for col in yr_cols], GREEN, USD0, italic=True); r+=1
row_h_ebit = r
set_hai_row(r, 'EBIT', [f'={col}{row_h_ebitda}-{col}{row_h_da}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_h_int = r
set_hai_row(r, 'Interest (shared — unaffected by Hosted.ai)', [f'=Income_Statement!{col}13' for col in yr_cols], GREEN, USD0, italic=True); r+=1
row_h_ebt = r
set_hai_row(r, 'EBT', [f'={col}{row_h_ebit}-{col}{row_h_int}' for col in yr_cols], BLACK, USD0); r+=1
row_h_tax = r
set_hai_row(r, 'Tax', [f'=MAX(0,{col}{row_h_ebt})*Assumptions!$C$35' for col in yr_cols], GREEN, USD0); r+=1
row_h_ni = r
set_hai_row(r, 'Net Income', [f'={col}{row_h_ebt}-{col}{row_h_tax}' for col in yr_cols], BLACK, USD0, bold=True); r+=1
row_h_fee_memo = r
set_hai_row(r, 'memo: Hosted.ai license fee (incl. above)', [f'={hai_fee_formula2.format(col=col)}' for col in yr_cols], GREEN, USD0, italic=True); r+=1

# ---- Cash flows & returns (columns B..G, B=Year0) ----
r += 1
style_section(hai, r, 2, 'PROJECT & EQUITY CASH FLOWS', span=6); r+=1
hdr_r = r
hai_ret_cols = ['B','C','D','E','F','G']
for i, col in enumerate(hai_ret_cols):
    c = hai[f'{col}{hdr_r}']; c.value = f'Year {i}'; c.font = BOLD_WHITE; c.fill = HEADER_FILL
r += 1

def set_hai_ret_row(row, label_text, formulas, bold=False):
    label(hai, row, 2, label_text, bold=bold)
    for col, f in zip(hai_ret_cols, formulas):
        cell = hai[f'{col}{row}']
        cell.value = f
        cell.font = Font(name='Arial', bold=bold, color='008000')
        cell.number_format = USD0

row_pfcf_w = r
formulas = ['=-Assumptions!C43']
for i, col in enumerate(yr_cols):
    yr_num = i+1
    base = f'{col}{row_w_ebit}*(1-Assumptions!$C$35)+{col}{row_w_da}'
    if yr_num==5: formulas.append(f'={base}+Assumptions!$C$43*Assumptions!$C$23')
    else: formulas.append(f'={base}')
set_hai_ret_row(r, 'Project FCF — Without Hosted.ai', formulas); r+=1

row_pfcf_h = r
formulas = ['=-Assumptions!C43']
for i, col in enumerate(yr_cols):
    yr_num = i+1
    base = f'{col}{row_h_ebit}*(1-Assumptions!$C$35)+{col}{row_h_da}'
    if yr_num==5: formulas.append(f'={base}+Assumptions!$C$43*Assumptions!$C$23')
    else: formulas.append(f'={base}')
set_hai_ret_row(r, 'Project FCF — With Hosted.ai', formulas); r+=1

row_efcf_w = r
formulas = ['=-Assumptions!C45']
for i, col in enumerate(yr_cols):
    yr_num = i+1
    base = f'{col}{row_w_ni}+{col}{row_w_da}+Cash_Flow!{col}8'
    if yr_num==5: formulas.append(f'={base}+Assumptions!$C$43*Assumptions!$C$23-Balance_Sheet!G11')
    else: formulas.append(f'={base}')
set_hai_ret_row(r, 'Equity FCF — Without Hosted.ai', formulas); r+=1

row_efcf_h = r
formulas = ['=-Assumptions!C45']
for i, col in enumerate(yr_cols):
    yr_num = i+1
    base = f'{col}{row_h_ni}+{col}{row_h_da}+Cash_Flow!{col}8'
    if yr_num==5: formulas.append(f'={base}+Assumptions!$C$43*Assumptions!$C$23-Balance_Sheet!G11')
    else: formulas.append(f'={base}')
set_hai_ret_row(r, 'Equity FCF — With Hosted.ai', formulas); r+=1

row_cumpfcf_w = r
set_hai_ret_row(r, 'Cumulative Project FCF — Without', [f'=SUM($B${row_pfcf_w}:{c}{row_pfcf_w})' for c in hai_ret_cols]); r+=1
row_cumpfcf_h = r
set_hai_ret_row(r, 'Cumulative Project FCF — With', [f'=SUM($B${row_pfcf_h}:{c}{row_pfcf_h})' for c in hai_ret_cols]); r+=1

r += 1
style_section(hai, r, 2, 'RETURNS & DELTA SUMMARY', span=6); r+=1

def set_metric(row, label_text, formula, numfmt=USD0, bold=True):
    label(hai, row, 2, label_text, bold=bold)
    c = hai.cell(row=row, column=3, value=formula)
    c.font = Font(name='Arial', bold=bold); c.number_format = numfmt

row_irr_pw = r; set_metric(r, 'Project IRR — Without', f'=IFERROR(IRR(B{row_pfcf_w}:G{row_pfcf_w}),"n/a")', PCT1); r+=1
row_irr_ph = r; set_metric(r, 'Project IRR — With', f'=IFERROR(IRR(B{row_pfcf_h}:G{row_pfcf_h}),"n/a")', PCT1); r+=1
row_irr_ew = r; set_metric(r, 'Equity IRR — Without', f'=IFERROR(IRR(B{row_efcf_w}:G{row_efcf_w},-0.5),"n/a")', PCT1); r+=1
row_irr_eh = r; set_metric(r, 'Equity IRR — With', f'=IFERROR(IRR(B{row_efcf_h}:G{row_efcf_h}),"n/a")', PCT1); r+=1
row_npv_pw = r; set_metric(r, 'NPV — Project, Without @ WACC', f'=B{row_pfcf_w}+NPV(Assumptions!C39,C{row_pfcf_w}:G{row_pfcf_w})'); r+=1
row_npv_ph = r; set_metric(r, 'NPV — Project, With @ WACC', f'=B{row_pfcf_h}+NPV(Assumptions!C39,C{row_pfcf_h}:G{row_pfcf_h})'); r+=1
row_npv_ew = r; set_metric(r, 'NPV — Equity, Without @ Cost of Equity', f'=B{row_efcf_w}+NPV(Assumptions!C40,C{row_efcf_w}:G{row_efcf_w})'); r+=1
row_npv_eh = r; set_metric(r, 'NPV — Equity, With @ Cost of Equity', f'=B{row_efcf_h}+NPV(Assumptions!C40,C{row_efcf_h}:G{row_efcf_h})'); r+=1
row_payback_w = r; set_metric(r, 'Simple payback — Without (break-even year)',
    f'=IFERROR(MATCH(1,INDEX((C{row_cumpfcf_w}:G{row_cumpfcf_w}>=0)*1,0),0),"Beyond horizon")', '0'); r+=1
row_payback_h = r; set_metric(r, 'Simple payback — With (break-even year)',
    f'=IFERROR(MATCH(1,INDEX((C{row_cumpfcf_h}:G{row_cumpfcf_h}>=0)*1,0),0),"Beyond horizon")', '0'); r+=1
r += 1
row_cum_rev_w = r; set_metric(r, 'Cumulative Revenue — Without (5yr)', f'=SUM(C{row_w_rev}:G{row_w_rev})'); r+=1
row_cum_rev_h = r; set_metric(r, 'Cumulative Revenue — With (5yr)', f'=SUM(C{row_h_rev}:G{row_h_rev})'); r+=1
row_cum_ebitda_w = r; set_metric(r, 'Cumulative EBITDA — Without (5yr)', f'=SUM(C{row_w_ebitda}:G{row_w_ebitda})'); r+=1
row_cum_ebitda_h = r; set_metric(r, 'Cumulative EBITDA — With (5yr)', f'=SUM(C{row_h_ebitda}:G{row_h_ebitda})'); r+=1
row_cum_fee = r; set_metric(r, 'Cumulative Hosted.ai license fees paid (5yr)', f'=SUM(C{row_h_fee_memo}:G{row_h_fee_memo})'); r+=1
r += 1
row_delta_npv = r; set_metric(r, 'Δ Project NPV (With − Without)', f'=C{row_npv_ph}-C{row_npv_pw}'); r+=1
row_roi_mult = r; set_metric(r, 'ROI multiple (ΔNPV ÷ cumulative license fee)', f'=IFERROR(C{row_delta_npv}/C{row_cum_fee},"n/a")', '0.0"x";"n/a"'); r+=1

for row in hai.iter_rows(min_row=3, max_row=r, min_col=2, max_col=7):
    for cell in row:
        cell.border = BORDER
hai.freeze_panes = 'C4'

a.freeze_panes = 'C4'
t.freeze_panes = 'C4'
inc.freeze_panes = 'C4'
bs.freeze_panes = 'C4'
cf.freeze_panes = 'C4'
ret.freeze_panes = 'C4'

tab_colors = {'Cover':'444444','Assumptions':'0B5394','TCO_BuildUp':'3ECF8E','Income_Statement':'4F8CFF',
              'Balance_Sheet':'4F8CFF','Cash_Flow':'4F8CFF','Returns':'FFB454','Hosted_ai_ROI':'B57BFF'}
for name, color in tab_colors.items():
    wb[name].sheet_properties.tabColor = color

wb.save('/sessions/zealous-elegant-volta/mnt/outputs/AI_Cloud_TCO_Model.xlsx')
print('Workbook built')
